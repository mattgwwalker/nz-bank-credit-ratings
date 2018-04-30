# Process the file from the Reserve Bank
# source: https://www.rbnz.govt.nz/statistics/g1
#
# The raw file from the Reserve Bank had to be manually edited to
# correct two clearly unintended dates of "June 31" for 2002, 2003,
# and 2012.  They were corrected to June 30.
#
# A second issue had to be manually corrected, where the Reserve Bank
# did not specify the year for TSB's September position in 2004.


import dateutil
import sys

# Scan through looking for bank names
def scanSheetForBanks(sheet):
    banks = []
    for x in range(1, sheet.max_row):
        d = sheet.cell(row=3, column=x)
        if d.value is not None:
            #print(d.value+"\tin column "+str(x))
            banks.append( (d.value, x) )
    return banks



# Attempt to process Excel date into a Python datetime object for
# storage in the database.
def processDate(d):
    try:
        processed = dateutil.parser.parse(d)
        return processed
    except TypeError:
        return d

# Function to get ratings for the banks (version 1)
def getRatings1(banks, periodEndingRow, ratingRow, issuerRow):
    result = []
    
    for (bankName,bankCol) in banks:
        if bankName==0: continue
        periodEnding = processDate(sheet.cell(row=periodEndingRow, column=bankCol).value)
        rating = sheet.cell(row=ratingRow, column=bankCol).value
        issuer = sheet.cell(row=issuerRow, column=bankCol).value
        if periodEnding is None: continue
        
        if rating!="None":
            try:
                ratings = rating.split("/")
                issuers = issuer.split("/")
                assert len(ratings) == len(issuers)
            except Exception:
                print(Exception)
                print("bankName",bankName)
                print("rating",rating)
                print("issuer", issuer)
                sys.exit()
            
            for r, i in zip(ratings,issuers):
                dataPoint = (bankName, periodEnding, r, i)
                #print(dataPoint)
                result.append(dataPoint)

    return result
        

# Function to get ratings for the banks (version 2)
def getRatings2(banks, periodEndingRow, standardAndPoorRow, moodyRow, fitchRow):
    ratings = []
    
    for (bankName,bankCol) in banks:
        if bankName==0: continue
        #print("bankCol", bankCol)
        #print("periodEndingRow",periodEndingRow)

        periodEnding = processDate(sheet.cell(row=periodEndingRow, column=bankCol).value)
        standardAndPoorRating = sheet.cell(row=standardAndPoorRow, column=bankCol).value
        moodyRating = sheet.cell(row=moodyRow, column=bankCol).value
        fitchRating = sheet.cell(row=fitchRow, column=bankCol).value

        if periodEnding is None: continue

        if standardAndPoorRating!="-" and standardAndPoorRating!="..":
            dataPoint = (bankName, periodEnding, standardAndPoorRating, "Standard and Poor's")
            ratings.append(dataPoint)

        if moodyRating!="-" and moodyRating!="..":
            dataPoint = (bankName, periodEnding, moodyRating, "Moody's")
            ratings.append(dataPoint)
            
        if fitchRating!="-" and fitchRating!="..":
            dataPoint = (bankName, periodEnding, fitchRating, "Fitch")
            ratings.append(dataPoint)

    return ratings


    

# Scan through looking for the credit rating
def scanSheetForRatings(sheet, banks):
    periodEndingRow = None
    ratingRow = None
    issuerRow = None
    standardAndPoorRow = None
    moodyRow = None
    fitchRow = None
    # Loop through cells in first column (and to some extent the second column too)
    for x in range(1, sheet.max_column):
        d = sheet.cell(row=x, column=1)
        e = sheet.cell(row=x, column=2)
        
        def isPeriodEnd(d):
            if d.value is None:
                return False
            try:
                result = ("End of period" in d.value and "disclosure statement" in d.value) or "For the period ending" in d.value
            except TypeError:
                result = False

            return result

        if isPeriodEnd(d) or isPeriodEnd(e):
            if periodEndingRow is not None:
                raise Exception("periodEndingRow has already been set")
            else:
                periodEndingRow = x


        if d.value=="Credit rating(s)":
            d1 = sheet.cell(row=x+1, column=1)
            d2 = sheet.cell(row=x+2, column=1)
            d3 = sheet.cell(row=x+3, column=1)
            d4 = sheet.cell(row=x+4, column=1)
            if d2.value is None or d3.value is None:
                d1 = sheet.cell(row=x+1, column=2)
                d2 = sheet.cell(row=x+2, column=2)
                d3 = sheet.cell(row=x+3, column=2)
                d4 = sheet.cell(row=x+4, column=2)
            #print("d2",d2.value)
            #print("d3",d3.value)
                
            if d2.value is not None and d3.value is not None:
                if "Rating" in d2.value and "Issuer" in d3.value:
                    if ratingRow is not None:
                        raise Exception("ratingRow has already been set")
                    else:
                        ratingRow = x+2
                        
                    if issuerRow is not None:
                        raise Exception("issuerRow has laready been set")
                    else:
                        issuerRow = x+3
                        
                elif "Standard and Poor" in d2.value and "Moody" in d3.value and "Fitch" in d4.value:
                    standardAndPoorRow = x+2
                    moodyRow = x+3
                    fitchRow = x+4
                    
                elif "Standard and Poor" in d1.value and "Moody" in d2.value and "Fitch" in d3.value:
                    standardAndPoorRow = x+1
                    moodyRow = x+2
                    fitchRow = x+3
                    
                else:
                    print("Debug values:")
                    print("d1",d1.value)
                    print("d2",d2.value)
                    print("d3",d3.value)
                    print("d4",d4.value)
                    
                
    # Go through all the banks and get their rating
    ratings = None
    if periodEndingRow is not None and ratingRow is not None and issuerRow is not None:
        ratings = getRatings1(banks, periodEndingRow, ratingRow, issuerRow)

    elif periodEndingRow is not None and standardAndPoorRow is not None and moodyRow is not None and fitchRow is not None:
        ratings = getRatings2(banks, periodEndingRow, standardAndPoorRow, moodyRow, fitchRow)

    else:
        print("periodEndingRow", periodEndingRow)
        print("ratingRow", ratingRow)
        print("issuerRow", issuerRow)
        print("standardAndPoorRow", standardAndPoorRow)
        print("moodyRow", moodyRow)
        print("fitchRow", fitchRow)
    return ratings




# ***************************************
# Read Excel file
# ***************************************


filename = "hg1.xlsx"

from openpyxl import load_workbook

print("Loading workbook")
wb = load_workbook(filename)

sheetNames = wb.sheetnames

ratings = []

print("Processing sheets")
for sheetName in sheetNames:
    #print("Processing sheet",sheetName)
    sheet = wb[sheetName]
    banks = scanSheetForBanks(sheet)
    sheetRatings = scanSheetForRatings(sheet, banks)
    if sheetRatings is not None:
        #print(sheetRatings)
        ratings.extend(sheetRatings)
    else:
        print("Error processing sheet", sheetName)


#print(ratings)
#print(len(ratings))



# ***************************************
# Write to SQLLite database
# ***************************************

print("Writing to SQL database")
import sqlite3

dbFile = "ratings.sqlite3"
conn = sqlite3.connect(dbFile)
c = conn.cursor()

for bankName, periodEnding, rating, issuer in ratings:
    c.execute("INSERT INTO SourceData (periodEnding, bankName, issuerName, ratingText) VALUES (?, ?, ?, ?)",
              (periodEnding,
               bankName,
               issuer,
               rating))

conn.commit()


# Bank names to bank IDs
# **********************

# ANZ
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('ANZ')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ Nat'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ National'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZN'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZN1'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ Bank Ltd.'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ Bank NZ Ltd.'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ANZ Bank NZ'", (bankId,))
conn.commit()

# ASB
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('ASB')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ASB'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ASB1'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ASB-BK'", (bankId,))
conn.commit()

# BNZ
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('BNZ')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='BNZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='BNZ1'", (bankId,))
conn.commit()

# Bankers Trust
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Bankers Trust')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Bankers Trust'", (bankId,))
conn.commit()

# Country-wide
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Country-wide')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Country-wide'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Countrywide'", (bankId,))
conn.commit()

# National
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('National')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='National'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='National Bank'", (bankId,))
conn.commit()

# Trust Bank
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Trust Bank')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Trust Bank'", (bankId,))
conn.commit()

# Rabo New Zealand
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Rabo New Zealand')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Rabo New Zealand'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Rabo NZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Rabo-NZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Rabobank'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='RABO-NZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Rabobank NZ'", (bankId,))
conn.commit()

# Kiwibank
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Kiwibank')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Kiwibank'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='KIWI'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Kiwi'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='KIWI1'", (bankId,))
conn.commit()

# TSB
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('TSB')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='TSB'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='TSB1'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='TSB Bank'", (bankId,))
conn.commit()

# St. George Bank
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('St. George Bank')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='St. George Bank NZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='St. George Bk. NZ'", (bankId,))
conn.commit()

# Westpac
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Westpac')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Westpac NZ Ltd'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Westpac NZ Ltd.'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Westpac NZ'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='WNZL'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='WNZL1'", (bankId,))
conn.commit()

# SBS
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('SBS')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='SBS Bank'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='SBS'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='SBS-Bk'", (bankId,))
conn.commit()

# Baroda
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Baroda')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Baroda'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='BARODA'", (bankId,))
conn.commit()

# Bank of India
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Bank of India')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Bank of India'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='BOI'", (bankId,))
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='BOI-NZ'", (bankId,))
conn.commit()

# Co-op Bank
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Co-op Bank')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Co-op Bank'", (bankId,))
conn.commit()

# Heartland
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Heartland')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Heartland'", (bankId,))
conn.commit()

# ICBC
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('ICBC')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='ICBC'", (bankId,))
conn.commit()

# China Construction
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('China Construction')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='China Construction'", (bankId,))
conn.commit()

# Bank of China
c.execute("INSERT INTO Banks (cannonicalName) VALUES ('Bank of China')")
bankId = c.lastrowid
c.execute("UPDATE SourceData SET bankId=? WHERE bankName='Bank of China'", (bankId,))
conn.commit()


# Issuer names and rating text to rating IDs
# *****************************************

# Standard and Poor's
c.execute("SELECT issuerId FROM Issuers WHERE cannonicalName='Standard and Poor''s'")
issuerId, = c.fetchone()
ratings = c.execute("SELECT ratingId, cannonicalText FROM Ratings WHERE issuerId=?", (issuerId,))
for ratingId, cannonicalText in ratings:
    c2 = conn.cursor()
    c2.execute("UPDATE SourceData SET ratingId=? WHERE issuerName='S&P' AND ratingText=?", (ratingId,cannonicalText))
    c2.execute("UPDATE SourceData SET ratingId=? WHERE issuerName='Standard and Poor''s' AND ratingText=?", (ratingId,cannonicalText))
conn.commit()    
    
# Moody's
c.execute("SELECT issuerId FROM Issuers WHERE cannonicalName='Moody''s'")
issuerId, = c.fetchone()
ratings = c.execute("SELECT ratingId, cannonicalText FROM Ratings WHERE issuerId=?", (issuerId,))
for ratingId, cannonicalText in ratings:
    c2 = conn.cursor()
    c2.execute("UPDATE SourceData SET ratingId=? WHERE issuerName='Moody''s' AND ratingText=?", (ratingId,cannonicalText))
    c2.execute("UPDATE SourceData SET ratingId=? WHERE issuerName='Moodys' AND ratingText=?", (ratingId,cannonicalText))
conn.commit()    

# Fitch
c.execute("SELECT issuerId FROM Issuers WHERE cannonicalName='Fitch'")
issuerId, = c.fetchone()
ratings = c.execute("SELECT ratingId, cannonicalText FROM Ratings WHERE issuerId=?", (issuerId,))
for ratingId, cannonicalText in ratings:
    c2 = conn.cursor()
    c2.execute("UPDATE SourceData SET ratingId=? WHERE issuerName='Fitch' AND ratingText=?", (ratingId,cannonicalText))
conn.commit()    

conn.close()
